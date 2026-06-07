from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import xlwings as xw
from openpyxl import load_workbook
from shutil import copy2


PROJECT_ROOT = Path(__file__).resolve().parents[2]
AUTO_PAYSLIP_DIR = PROJECT_ROOT / "docs" / "Auto-Payslip"
DEFAULT_CALC_PATH = AUTO_PAYSLIP_DIR / "TNBGAJICALCULATION.xlsx"
DEFAULT_MASTER_PATH = AUTO_PAYSLIP_DIR / "CONSTRUCTION TEAM BSG.xlsx"
DEFAULT_PAYSLIP_TEMPLATE_PATH = AUTO_PAYSLIP_DIR / "Masburan Salary Template.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "results" / "payslips"
DEFAULT_LKS_SAMPLE_PATH = AUTO_PAYSLIP_DIR / "LKS Sample.xlsm"

SUPERVISOR_CODE = "KMRS0001"
FIXED_DEDUCTION_TOTAL = 180.0
VALID_TEAM_CODES = tuple(f"KMRT{index:04d}" for index in range(1, 9))
TEAM_CODE_SET = set(VALID_TEAM_CODES)

COUNT_CELL_ORDER = ("C", "D", "E", "F", "G", "H")
COUNT_TEMPLATE_ROWS = (14, 15, 16, 17, 18, 19)
COUNT_TEMPLATE_RATE_CELLS = ("E14", "E15", "E16", "E17", "E18", "E19")
KIV_TEMPLATE_ROWS = (21, 22)
KIV_TEMPLATE_RATE_CELLS = ("E21", "E22")
KIV_TEMPLATE_UNIT_CELLS = ("F21", "F22")
KIV_TEMPLATE_AMOUNT_CELLS = ("G21", "G22")
SUPERVISOR_LABEL_CELL = "D14"
SUPERVISOR_AMOUNT_CELL = "G14"

MAIN_INPUT_ROW_START = 5
MAIN_INPUT_ROW_END = 12
MAIN_GROSS_ROW_START = 19
MAIN_GROSS_ROW_END = 25
FINAL_ROW_START = 46
FINAL_ROW_END = 52
SUPERVISOR_ROW = 53

CLAIM_REQUIRED_HEADERS = {
    "Labor": "labor",
    "Voltage": "voltage",
    "Hari Biasa / Hujung Minggu / Cuti Umum": "day_type",
}
CLAIM_OPTIONAL_HEADERS = {
    "REMARKS 2": "remarks_2",
}
DAY_TYPE_TO_COLUMN = {
    ("HARI BIASA", "PH1"): "C",
    ("HARI BIASA", "PH3"): "D",
    ("HUJUNG MINGGU", "PH1"): "E",
    ("HUJUNG MINGGU", "PH3"): "F",
    ("CUTI UMUM", "PH1"): "G",
    ("CUTI UMUM", "PH3"): "H",
}
DAY_TYPE_LABELS = {
    "HARI BIASA": "Hari Biasa",
    "HUJUNG MINGGU": "Hujung Minggu",
    "WEEKEND": "Hujung Minggu",
    "CUTI UMUM": "Cuti Umum",
}
VOLTAGE_TO_PHASE = {"01": "PH1", "1": "PH1", "02": "PH3", "2": "PH3"}
ROLE_RATES = {
    "helper": (12.0, 14.5, 15.5, 16.5, 17.0, 18.5),
    "installer": (15.0, 17.5, 19.5, 20.5, 22.0, 23.5),
}
KIV_ROLE_RATES = {
    "helper": (15.0, 18.0),
    "installer": (18.0, 20.0),
}
POSITION_LABELS = {
    "helper": "HELPER",
    "installer": "WIREMAN PW4",
    "supervisor": "SUPERVISOR",
}


@dataclass(frozen=True)
class WorkerIdentity:
    worker_code: str
    team_code: str
    supervisor_code: str
    name: str
    ic_number: str
    position: str
    role: str


@dataclass(frozen=True)
class TeamCalculation:
    team_code: str
    counts: tuple[float, float, float, float, float, float]
    kiv_counts: tuple[float, float]
    helper_base_gross: float
    installer_base_gross: float
    helper_final_gross: float
    installer_final_gross: float
    helper_final_net: float
    installer_final_net: float


@dataclass(frozen=True)
class SupervisorCalculation:
    gross: float
    net: float


@dataclass(frozen=True)
class PayslipEntry:
    role: str
    worker_code: str
    team_code: str
    name: str
    ic_number: str
    position: str
    salary_month: str
    payment_date: date
    counts: tuple[float, float, float, float, float, float]
    kiv_counts: tuple[float, float]
    gross: float
    net: float
    deduction_total: float
    task_force_add_on: float
    template_path: Path


@dataclass
class GeneratedPayslip:
    entry: PayslipEntry
    xlsx_path: Path
    pdf_path: Path


@dataclass(frozen=True)
class FileClaimSummary:
    file_name: str
    total_rows: int
    counted_rows: int
    skipped_rows: int
    counts_by_team: dict[str, tuple[float, float, float, float, float, float]]
    kiv_counts_by_team: dict[str, tuple[float, float]]


@dataclass(frozen=True)
class ClaimCountSummary:
    source_files: int
    total_rows: int
    counted_rows: int
    skipped_rows: int
    counts_by_team: dict[str, tuple[float, float, float, float, float, float]]
    kiv_counts_by_team: dict[str, tuple[float, float]]
    file_summaries: list[FileClaimSummary]
    warnings: list[str]


@dataclass
class PayslipGenerationResult:
    output_dir: Path
    generated: list[GeneratedPayslip] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    pdf_failures: list[str] = field(default_factory=list)
    calculation_workbook_path: Path | None = None
    claim_summary: ClaimCountSummary | None = None

    @property
    def generated_xlsx_count(self) -> int:
        return len(self.generated)

    @property
    def generated_pdf_count(self) -> int:
        return len(self.generated) - len(self.pdf_failures)


def _as_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_float(value: object) -> float:
    if value is None or value == "":
        return 0.0
    return float(value)


def _normalize_role(position: str) -> str | None:
    upper = position.upper()
    if "HELPER" in upper or "PEMBANTU" in upper:
        return "helper"
    if "INSTALLER" in upper or "PEMASANG" in upper:
        return "installer"
    if "SUPERVISOR" in upper or "PENYELIA" in upper:
        return "supervisor"
    return None


def _sanitize_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._ -]+", "", value).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "untitled"


def _normalize_header_text(value: object) -> str:
    text = _as_text(value)
    text = re.sub(r"\s+", " ", text)
    return text.strip().upper()


def _normalize_team_code(value: str) -> str | None:
    upper = value.strip().upper()
    match = re.fullmatch(r"[ZK]MRT(\d{4})", upper)
    if not match:
        return None
    team_code = f"KMRT{match.group(1)}"
    if team_code not in TEAM_CODE_SET:
        return None
    return team_code


def _normalize_phase(value: object) -> str | None:
    text = _as_text(value)
    return VOLTAGE_TO_PHASE.get(text)


def _normalize_day_type(value: object) -> str | None:
    text = _as_text(value).upper()
    if not text:
        return None
    return DAY_TYPE_LABELS.get(text)


def _format_month_slug(payment_date: date, salary_month: str) -> str:
    month_label = salary_month.strip() or payment_date.strftime("%B %Y")
    return _sanitize_filename(month_label)


def _build_output_root(output_dir: Path, payment_date: date, salary_month: str) -> Path:
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    month_slug = _format_month_slug(payment_date, salary_month)
    return output_dir / month_slug / timestamp


def load_worker_master(master_path: Path) -> tuple[dict[str, dict[str, WorkerIdentity]], WorkerIdentity | None]:
    workbook = load_workbook(master_path, read_only=True, data_only=True)
    try:
        worksheet = workbook["DATA PERSONAL"]

        team_members: dict[str, dict[str, WorkerIdentity]] = {}
        supervisor: WorkerIdentity | None = None

        for row in worksheet.iter_rows(min_row=1, values_only=True):
            team_code = _as_text(row[2] if len(row) > 2 else None)
            supervisor_code = _as_text(row[3] if len(row) > 3 else None)
            name = _as_text(row[4] if len(row) > 4 else None)
            ic_number = _as_text(row[5] if len(row) > 5 else None)
            position = _as_text(row[6] if len(row) > 6 else None)

            if not team_code or not name or not position:
                continue

            role = _normalize_role(position)
            if role is None:
                continue

            identity = WorkerIdentity(
                worker_code=team_code if role != "supervisor" else supervisor_code or team_code,
                team_code=team_code,
                supervisor_code=supervisor_code,
                name=name,
                ic_number=ic_number,
                position=position,
                role=role,
            )

            if role == "supervisor":
                if identity.worker_code == SUPERVISOR_CODE or supervisor is None:
                    supervisor = identity
                continue

            team_members.setdefault(team_code, {})[role] = identity

        return team_members, supervisor
    finally:
        workbook.close()


def _find_claim_header_row(worksheet) -> tuple[int, dict[str, int]]:
    required_headers = {
        _normalize_header_text(header): key for header, key in CLAIM_REQUIRED_HEADERS.items()
    }
    optional_headers = {
        _normalize_header_text(header): key for header, key in CLAIM_OPTIONAL_HEADERS.items()
    }

    for row_idx in range(1, min(10, worksheet.max_row) + 1):
        column_map: dict[str, int] = {}
        for col_idx in range(1, worksheet.max_column + 1):
            header = _normalize_header_text(worksheet.cell(row=row_idx, column=col_idx).value)
            key = required_headers.get(header)
            if key is not None:
                column_map[key] = col_idx
                continue
            optional_key = optional_headers.get(header)
            if optional_key is not None:
                column_map[optional_key] = col_idx
        required_keys = set(CLAIM_REQUIRED_HEADERS.values())
        if required_keys.issubset(column_map.keys()):
            return row_idx, column_map

    raise ValueError("Could not find the required CLAIM headers in the selected LKS workbook.")


def load_claim_counts(lks_paths: list[Path]) -> ClaimCountSummary:
    counts: dict[str, list[float]] = {team_code: [0.0] * 6 for team_code in VALID_TEAM_CODES}
    kiv_counts: dict[str, list[float]] = {team_code: [0.0, 0.0] for team_code in VALID_TEAM_CODES}
    file_summaries: list[FileClaimSummary] = []
    warnings: list[str] = []
    total_rows = 0
    counted_rows = 0

    for lks_path in lks_paths:
        file_counts: dict[str, list[float]] = {team_code: [0.0] * 6 for team_code in VALID_TEAM_CODES}
        file_kiv_counts: dict[str, list[float]] = {team_code: [0.0, 0.0] for team_code in VALID_TEAM_CODES}
        file_total_rows = 0
        file_counted_rows = 0
        workbook = load_workbook(lks_path, read_only=True, data_only=True)
        try:
            if "CLAIM" not in workbook.sheetnames:
                warnings.append(f"{lks_path.name}: missing CLAIM sheet.")
                continue

            worksheet = workbook["CLAIM"]
            header_row, column_map = _find_claim_header_row(worksheet)

            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                team_code = _normalize_team_code(worksheet.cell(row=row_idx, column=column_map["labor"]).value or "")
                phase = _normalize_phase(worksheet.cell(row=row_idx, column=column_map["voltage"]).value)
                day_type = _normalize_day_type(worksheet.cell(row=row_idx, column=column_map["day_type"]).value)
                remarks_2 = _as_text(
                    worksheet.cell(row=row_idx, column=column_map.get("remarks_2", 0)).value
                ).upper() if column_map.get("remarks_2") else ""
                is_kiv = "KIV" in remarks_2

                total_rows += 1
                file_total_rows += 1
                if team_code is None or phase is None:
                    continue

                if is_kiv:
                    kiv_index = 0 if phase == "PH1" else 1
                    kiv_counts[team_code][kiv_index] += 1.0
                    file_kiv_counts[team_code][kiv_index] += 1.0
                else:
                    if day_type is None:
                        continue
                    column = DAY_TYPE_TO_COLUMN.get((day_type.upper(), phase))
                    if column is None:
                        continue

                    count_index = COUNT_CELL_ORDER.index(column)
                    counts[team_code][count_index] += 1.0
                    file_counts[team_code][count_index] += 1.0
                counted_rows += 1
                file_counted_rows += 1
        finally:
            workbook.close()

        file_summaries.append(
            FileClaimSummary(
                file_name=lks_path.name,
                total_rows=file_total_rows,
                counted_rows=file_counted_rows,
                skipped_rows=file_total_rows - file_counted_rows,
                counts_by_team={team_code: tuple(values) for team_code, values in file_counts.items()},
                kiv_counts_by_team={team_code: tuple(values) for team_code, values in file_kiv_counts.items()},
            )
        )

    return ClaimCountSummary(
        source_files=len(lks_paths),
        total_rows=total_rows,
        counted_rows=counted_rows,
        skipped_rows=total_rows - counted_rows,
        counts_by_team={team_code: tuple(values) for team_code, values in counts.items()},
        kiv_counts_by_team={team_code: tuple(values) for team_code, values in kiv_counts.items()},
        file_summaries=file_summaries,
        warnings=warnings + (["No valid CLAIM rows were counted from the selected LKS files."] if counted_rows == 0 else []),
    )


def create_calculation_workbook(
    template_path: Path,
    output_dir: Path,
    salary_month: str,
    payment_date: date,
    lks_paths: list[Path],
) -> tuple[Path, ClaimCountSummary]:
    claim_summary = load_claim_counts(lks_paths)
    workbook = load_workbook(template_path)
    try:
        worksheet = workbook[workbook.sheetnames[0]]

        row_by_team: dict[str, int] = {}
        for row_idx in range(MAIN_INPUT_ROW_START, MAIN_INPUT_ROW_END + 1):
            team_code = _normalize_team_code(_as_text(worksheet[f"B{row_idx}"].value) or "")
            if team_code:
                row_by_team[team_code] = row_idx

        kiv_row_by_team: dict[str, int] = {}
        for row_idx in range(33, 41):
            team_code = _normalize_team_code(_as_text(worksheet[f"B{row_idx}"].value) or "")
            if team_code:
                kiv_row_by_team[team_code] = row_idx

        for team_code in VALID_TEAM_CODES:
            row_idx = row_by_team.get(team_code)
            if row_idx is None:
                continue

            values = claim_summary.counts_by_team.get(team_code, (0.0, 0.0, 0.0, 0.0, 0.0, 0.0))
            for column, value in zip(COUNT_CELL_ORDER, values):
                worksheet[f"{column}{row_idx}"] = value

            kiv_row_idx = kiv_row_by_team.get(team_code)
            if kiv_row_idx is not None:
                kiv_ph1, kiv_ph3 = claim_summary.kiv_counts_by_team.get(team_code, (0.0, 0.0))
                worksheet[f"C{kiv_row_idx}"] = kiv_ph1
                worksheet[f"D{kiv_row_idx}"] = kiv_ph3
                worksheet[f"E{kiv_row_idx}"] = 0
                worksheet[f"F{kiv_row_idx}"] = 0
                worksheet[f"G{kiv_row_idx}"] = 0
                worksheet[f"H{kiv_row_idx}"] = 0

        if hasattr(workbook, "calculation") and workbook.calculation is not None:
            workbook.calculation.calcMode = "auto"
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True

        calc_dir = output_dir / "calculation"
        calc_dir.mkdir(parents=True, exist_ok=True)
        month_slug = _format_month_slug(payment_date, salary_month)
        output_path = calc_dir / f"{month_slug}_TNBGAJICALCULATION.xlsx"
        workbook.save(output_path)
        return output_path, claim_summary
    finally:
        workbook.close()


def format_claim_summary_lines(claim_summary: ClaimCountSummary) -> list[str]:
    lines = [
        f"LKS files used: {claim_summary.source_files}",
        f"Total CLAIM rows: {claim_summary.total_rows}",
        f"Counted CLAIM rows: {claim_summary.counted_rows}",
        f"Skipped CLAIM rows: {claim_summary.skipped_rows}",
    ]

    for file_summary in claim_summary.file_summaries:
        lines.append("")
        lines.append(f"{file_summary.file_name}")
        lines.append(f"Counted rows: {file_summary.counted_rows}/{file_summary.total_rows}")
        if file_summary.skipped_rows:
            lines.append(f"  Skipped rows: {file_summary.skipped_rows}")
        lines.extend(_format_team_count_lines(file_summary.counts_by_team))
        if any(any(values) for values in file_summary.kiv_counts_by_team.values()):
            lines.append("KIV totals:")
            lines.extend(_format_kiv_count_lines(file_summary.kiv_counts_by_team))

    lines.append("")
    lines.append("Combined team totals:")
    lines.extend(_format_team_count_lines(claim_summary.counts_by_team))
    if any(any(values) for values in claim_summary.kiv_counts_by_team.values()):
        lines.append("Combined KIV totals:")
        lines.extend(_format_kiv_count_lines(claim_summary.kiv_counts_by_team))
    return lines


def _format_team_count_lines(
    counts_by_team: dict[str, tuple[float, float, float, float, float, float]],
) -> list[str]:
    lines = [
        "TEAM     TOTAL  HB-P1  HB-P3  HM-P1  HM-P3  CU-P1  CU-P3",
        "-------  -----  -----  -----  -----  -----  -----  -----",
    ]
    for team_code in VALID_TEAM_CODES:
        values = counts_by_team.get(team_code, (0.0, 0.0, 0.0, 0.0, 0.0, 0.0))
        if not any(values):
            continue
        hb_ph1, hb_ph3, hm_ph1, hm_ph3, cu_ph1, cu_ph3 = values
        total = int(sum(values))
        lines.append(
            f"{team_code:<7}  {total:>5}  {int(hb_ph1):>5}  {int(hb_ph3):>5}  "
            f"{int(hm_ph1):>5}  {int(hm_ph3):>5}  {int(cu_ph1):>5}  {int(cu_ph3):>5}"
        )
    return lines


def _format_kiv_count_lines(
    counts_by_team: dict[str, tuple[float, float]],
) -> list[str]:
    lines = [
        "TEAM     TOTAL  KIV-P1  KIV-P3",
        "-------  -----  ------  ------",
    ]
    for team_code in VALID_TEAM_CODES:
        kiv_ph1, kiv_ph3 = counts_by_team.get(team_code, (0.0, 0.0))
        if not kiv_ph1 and not kiv_ph3:
            continue
        total = int(kiv_ph1 + kiv_ph3)
        lines.append(f"{team_code:<7}  {total:>5}  {int(kiv_ph1):>6}  {int(kiv_ph3):>6}")
    return lines


def recalculate_workbook(workbook_path: Path) -> None:
    app: xw.App | None = None
    book = None
    try:
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        book = app.books.open(str(workbook_path), update_links=False, read_only=False)
        app.api.CalculateFull()
        book.save()
    finally:
        if book is not None:
            book.close()
        if app is not None:
            app.quit()


def load_calculation(calc_path: Path) -> tuple[list[TeamCalculation], SupervisorCalculation]:
    workbook = load_workbook(calc_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]

        counts_by_team: dict[str, tuple[float, float, float, float, float, float]] = {}
        for row_idx in range(MAIN_INPUT_ROW_START, MAIN_INPUT_ROW_END + 1):
            team_code = _as_text(worksheet[f"B{row_idx}"].value)
            if not team_code:
                continue
            counts_by_team[team_code] = tuple(
                _as_float(worksheet[f"{column}{row_idx}"].value) for column in COUNT_CELL_ORDER
            )

        kiv_counts_by_team: dict[str, tuple[float, float]] = {}
        for row_idx in range(33, 41):
            team_code = _as_text(worksheet[f"B{row_idx}"].value)
            if not team_code:
                continue
            kiv_counts_by_team[team_code] = (
                _as_float(worksheet[f"C{row_idx}"].value),
                _as_float(worksheet[f"D{row_idx}"].value),
            )

        base_gross_by_team: dict[str, tuple[float, float]] = {}
        for row_idx in range(MAIN_GROSS_ROW_START, MAIN_GROSS_ROW_END + 1):
            team_code = _as_text(worksheet[f"B{row_idx}"].value)
            if not team_code:
                continue
            base_gross_by_team[team_code] = (
                _as_float(worksheet[f"C{row_idx}"].value),
                _as_float(worksheet[f"E{row_idx}"].value),
            )

        team_calculations: list[TeamCalculation] = []
        for row_idx in range(FINAL_ROW_START, FINAL_ROW_END + 1):
            team_code = _as_text(worksheet[f"I{row_idx}"].value)
            if not team_code:
                continue

            counts = counts_by_team.get(team_code, (0.0, 0.0, 0.0, 0.0, 0.0, 0.0))
            helper_base_gross, installer_base_gross = base_gross_by_team.get(team_code, (0.0, 0.0))

            team_calculations.append(
                TeamCalculation(
                    team_code=team_code,
                    counts=counts,
                    kiv_counts=kiv_counts_by_team.get(team_code, (0.0, 0.0)),
                    helper_base_gross=helper_base_gross,
                    installer_base_gross=installer_base_gross,
                    helper_final_gross=_as_float(worksheet[f"J{row_idx}"].value),
                    installer_final_gross=_as_float(worksheet[f"L{row_idx}"].value),
                    helper_final_net=_as_float(worksheet[f"P{row_idx}"].value),
                    installer_final_net=_as_float(worksheet[f"R{row_idx}"].value),
                )
            )

        supervisor = SupervisorCalculation(
            gross=_as_float(worksheet[f"J{SUPERVISOR_ROW}"].value),
            net=_as_float(worksheet[f"P{SUPERVISOR_ROW}"].value),
        )
        return team_calculations, supervisor
    finally:
        workbook.close()


def build_entries(
    calculations: list[TeamCalculation],
    supervisor_calc: SupervisorCalculation,
    team_members: dict[str, dict[str, WorkerIdentity]],
    supervisor: WorkerIdentity | None,
    salary_month: str,
    payment_date: date,
) -> tuple[list[PayslipEntry], list[str]]:
    entries: list[PayslipEntry] = []
    warnings: list[str] = []

    for calc in calculations:
        workers = team_members.get(calc.team_code)
        if not workers:
            warnings.append(f"Missing worker mapping for team {calc.team_code}.")
            continue

        helper = workers.get("helper")
        installer = workers.get("installer")

        if helper is None:
            warnings.append(f"Missing helper mapping for team {calc.team_code}.")
        else:
            entries.append(
                PayslipEntry(
                    role="helper",
                    worker_code=helper.worker_code,
                    team_code=calc.team_code,
                    name=helper.name,
                    ic_number=helper.ic_number,
                    position=helper.position,
                    salary_month=salary_month,
                    payment_date=payment_date,
                    counts=calc.counts,
                    kiv_counts=calc.kiv_counts,
                    gross=calc.helper_final_gross,
                    net=calc.helper_final_net,
                    deduction_total=FIXED_DEDUCTION_TOTAL,
                    task_force_add_on=calc.helper_final_gross - calc.helper_base_gross,
                    template_path=DEFAULT_PAYSLIP_TEMPLATE_PATH,
                )
            )

        if installer is None:
            warnings.append(f"Missing installer mapping for team {calc.team_code}.")
        else:
            entries.append(
                PayslipEntry(
                    role="installer",
                    worker_code=installer.worker_code,
                    team_code=calc.team_code,
                    name=installer.name,
                    ic_number=installer.ic_number,
                    position=installer.position,
                    salary_month=salary_month,
                    payment_date=payment_date,
                    counts=calc.counts,
                    kiv_counts=calc.kiv_counts,
                    gross=calc.installer_final_gross,
                    net=calc.installer_final_net,
                    deduction_total=FIXED_DEDUCTION_TOTAL,
                    task_force_add_on=calc.installer_final_gross - calc.installer_base_gross,
                    template_path=DEFAULT_PAYSLIP_TEMPLATE_PATH,
                )
            )

    if supervisor is None:
        warnings.append(f"Missing supervisor mapping for {SUPERVISOR_CODE}.")
    else:
        entries.append(
            PayslipEntry(
                role="supervisor",
                worker_code=supervisor.worker_code,
                team_code=supervisor.team_code,
                name=supervisor.name,
                ic_number=supervisor.ic_number,
                position=supervisor.position,
                salary_month=salary_month,
                payment_date=payment_date,
                counts=(0.0, 0.0, 0.0, 0.0, 0.0, 0.0),
                kiv_counts=(0.0, 0.0),
                gross=supervisor_calc.gross,
                net=supervisor_calc.net,
                deduction_total=FIXED_DEDUCTION_TOTAL,
                task_force_add_on=0.0,
                template_path=DEFAULT_PAYSLIP_TEMPLATE_PATH,
            )
        )

    return entries, warnings


def _set_header_fields(worksheet, entry: PayslipEntry) -> None:
    worksheet["D9"] = entry.name
    worksheet["J9"] = entry.salary_month
    worksheet["D10"] = entry.ic_number
    worksheet["J10"] = entry.payment_date
    worksheet["D11"] = POSITION_LABELS.get(entry.role, entry.position)


def _clear_range(worksheet, addresses: list[str]) -> None:
    for address in addresses:
        worksheet[address] = None


def populate_payslip_template(entry: PayslipEntry, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    copy2(entry.template_path, output_path)

    app: xw.App | None = None
    book = None
    try:
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        book = app.books.open(str(output_path), update_links=False, read_only=False)
        sheet = book.sheets[0]

        sheet["D9"].value = entry.name
        sheet["J9"].value = entry.salary_month
        sheet["D10"].value = entry.ic_number
        sheet["J10"].value = entry.payment_date
        sheet["D11"].value = POSITION_LABELS.get(entry.role, entry.position)

        if entry.role in {"helper", "installer"}:
            for cell, rate in zip(COUNT_TEMPLATE_RATE_CELLS, ROLE_RATES[entry.role]):
                sheet[cell].value = rate
            for row, value in zip(COUNT_TEMPLATE_ROWS, entry.counts):
                sheet[f"F{row}"].value = value

            kiv_rates = KIV_ROLE_RATES[entry.role]
            kiv_counts = entry.kiv_counts
            for cell in ("D20", "E20", "F20", "G20"):
                sheet[cell].value = None
            sheet["D21"].value = "TASK FORCE / KIV (PH1)"
            sheet["D22"].value = "TASK FORCE / KIV (PH3)"
            for cell, rate in zip(KIV_TEMPLATE_RATE_CELLS, kiv_rates):
                sheet[cell].value = rate
            for cell, value in zip(KIV_TEMPLATE_UNIT_CELLS, kiv_counts):
                sheet[cell].value = value
            for cell, rate_cell, unit_cell in zip(KIV_TEMPLATE_AMOUNT_CELLS, KIV_TEMPLATE_RATE_CELLS, KIV_TEMPLATE_UNIT_CELLS):
                sheet[cell].formula = f"={rate_cell}*{unit_cell}"

            sheet["D23"].value = ""
            sheet["G23"].value = None
            for cell in ("J14", "J15", "J16", "J17"):
                sheet[cell].value = None
            sheet["J14"].value = entry.deduction_total
        else:
            sheet[SUPERVISOR_LABEL_CELL].value = "ALLOWANCE"
            sheet[SUPERVISOR_AMOUNT_CELL].value = entry.gross
            for cell in (
                "E14", "E15", "E16", "E17", "E18", "E19",
                "F14", "F15", "F16", "F17", "F18", "F19",
                "D20", "D21", "D22", "E20", "E21", "E22", "F20", "F21", "F22", "G20", "G21", "G22", "D23", "G23",
            ):
                sheet[cell].value = None
            for cell in ("J14", "J15", "J16", "J17"):
                sheet[cell].value = None
            sheet["J14"].value = entry.deduction_total

        app.api.CalculateFull()
        book.save()
    finally:
        if book is not None:
            book.close()
        if app is not None:
            app.quit()


def export_pdfs(generated: list[GeneratedPayslip]) -> list[str]:
    if not generated:
        return []

    failures: list[str] = []
    app: xw.App | None = None
    try:
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False

        for item in generated:
            book = None
            try:
                book = app.books.open(str(item.xlsx_path), update_links=False, read_only=False)
                app.api.CalculateFull()
                book.save()
                book.api.ExportAsFixedFormat(0, str(item.pdf_path))
            except Exception as exc:  # pragma: no cover - runtime only on Windows+Excel
                failures.append(f"{item.xlsx_path.name}: {exc}")
            finally:
                if book is not None:
                    book.close()
    finally:
        if app is not None:
            app.quit()

    return failures


def build_output_name(entry: PayslipEntry, payment_date: date) -> str:
    month_part = payment_date.strftime("%Y-%m")
    name_part = _sanitize_filename(entry.name)
    role_part = entry.role.upper()
    return f"{month_part}_{entry.team_code}_{role_part}_{name_part}"


def generate_payslips(
    calc_path: Path,
    master_path: Path,
    output_dir: Path,
    salary_month: str,
    payment_date: date,
    lks_paths: list[Path] | None = None,
) -> PayslipGenerationResult:
    run_output_dir = _build_output_root(output_dir, payment_date, salary_month)

    claim_summary: ClaimCountSummary | None = None
    effective_calc_path = calc_path
    if lks_paths:
        effective_calc_path, claim_summary = create_calculation_workbook(
            template_path=calc_path,
            output_dir=run_output_dir,
            salary_month=salary_month,
            payment_date=payment_date,
            lks_paths=lks_paths,
        )
        recalculate_workbook(effective_calc_path)

    team_members, supervisor = load_worker_master(master_path)
    calculations, supervisor_calc = load_calculation(effective_calc_path)
    entries, warnings = build_entries(
        calculations=calculations,
        supervisor_calc=supervisor_calc,
        team_members=team_members,
        supervisor=supervisor,
        salary_month=salary_month,
        payment_date=payment_date,
    )
    if claim_summary is not None:
        warnings = [*claim_summary.warnings, *warnings]

    excel_dir = run_output_dir / "excel"
    pdf_dir = run_output_dir / "pdf"

    generated: list[GeneratedPayslip] = []
    for entry in entries:
        output_name = build_output_name(entry, payment_date)
        xlsx_path = excel_dir / f"{output_name}.xlsx"
        pdf_path = pdf_dir / f"{output_name}.pdf"
        populate_payslip_template(entry, xlsx_path)
        pdf_path.parent.mkdir(parents=True, exist_ok=True)
        generated.append(GeneratedPayslip(entry=entry, xlsx_path=xlsx_path, pdf_path=pdf_path))

    pdf_failures = export_pdfs(generated)

    return PayslipGenerationResult(
        output_dir=run_output_dir,
        generated=generated,
        warnings=warnings,
        pdf_failures=pdf_failures,
        calculation_workbook_path=effective_calc_path if lks_paths else None,
        claim_summary=claim_summary,
    )
