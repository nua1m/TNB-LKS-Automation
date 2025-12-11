from openpyxl.styles import PatternFill, Alignment
from core.so_utils import clean_so
from config import ATTACH_SHEET_NAME, CLAIM_SHEET_NAME

RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

class QualityControl:
    @staticmethod
    def analyze_missing(handler):
        """Identifies SOs with missing images in Attachment Sheet."""
        wsA = handler.ws_attach
        missing_detail = {}
        counts = {"old": 0, "card": 0, "new": 0}
        
        col_old, col_card, col_new = 4, 5, 6
        
        for r in range(3, wsA.max_row + 1):
            so = clean_so(wsA.cell(r, 2).value)
            if not so: continue

            slots = []
            if not wsA.cell(r, col_old).value:
                slots.append("old_meter")
                counts["old"] += 1
            if not wsA.cell(r, col_card).value:
                slots.append("card")
                counts["card"] += 1
            if not wsA.cell(r, col_new).value:
                slots.append("new_meter")
                counts["new"] += 1
            
            if slots:
                missing_detail[so] = slots

        return missing_detail, counts

    @staticmethod
    def mark_defective(handler, missing_detail):
        """Highlights defective rows in RED."""
        defective_set = set(missing_detail.keys())
        wsC = handler.ws_claim
        wsA = handler.ws_attach

        # Helper to highlight row
        def highlight_row(ws, r):
            for cell in ws[r]:
                cell.fill = RED_FILL

        # CLAIM
        for r in range(3, wsC.max_row + 1):
            if clean_so(wsC.cell(r, 2).value) in defective_set:
                highlight_row(wsC, r)

        # ATTACHMENT
        for r in range(3, wsA.max_row + 1):
            if clean_so(wsA.cell(r, 2).value) in defective_set:
                highlight_row(wsA, r)

    @staticmethod
    def format_all(handler):
        """Applies center alignment to all cells."""
        for ws in [handler.ws_claim, handler.ws_attach]:
            for r in range(3, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).alignment = CENTER
