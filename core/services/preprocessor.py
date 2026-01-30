import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

class Preprocessor:
    @staticmethod
    def process_legacy_file(file_path):
        """
        New Workflow:
        1. Convert .xls -> temp.xlsx
        2. Open temp.xlsx
        3. Extract Date (Rows 1-14)
        4. Clean: Delete Top 14, Col D, Col T, Footer. Unmerge. Resize.
        5. Save as 'LKS Data (<Date>).xlsx'
        6. Return new path.
        """
        import win32com.client as win32
        import shutil
        import os
        from pathlib import Path
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        import re
        
        file_path = Path(file_path).resolve()
        temp_xlsx = file_path.with_name(file_path.stem + "_temp.xlsx")
        
        # 1. Convert to XLSX (Win32Com)
        try:
            # Use Dispatch 
            excel = win32.Dispatch('Excel.Application')
            excel.Visible = False # Run in background
            excel.DisplayAlerts = False
            
            # CorruptLoad=1 for robustness
            wb = excel.Workbooks.Open(str(file_path), UpdateLinks=0, CorruptLoad=1)
            wb.SaveAs(str(temp_xlsx), FileFormat=51) # xlOpenXMLWorkbook
            wb.Close(SaveChanges=False)
        except Exception as e:
            print(f"  Error converting legacy file: {e}")
            raise e
        finally:
            try: excel.Quit()
            except: pass

        # 2. Open with OpenPyXL
        try:
            wb = load_workbook(temp_xlsx)
            ws = wb.active
            
            # 3. Extract Date from Metadata (Rows 11 and 12)
            # User request: "take first date from row 11 and last date from row 12"
            # Format: (Jan 26 - Jan 30)
            
            from datetime import datetime
            import dateutil.parser
            
            def find_date_in_row(row_idx):
                for c in range(1, 20): # Scan first 20 cols
                    val = str(ws.cell(row_idx, c).value).strip()
                    if not val or val == "None": continue
                    
                    # Try fuzzy parse
                    try:
                        # Skip short random numbers or text
                        if len(val) < 6: continue 
                        dt = dateutil.parser.parse(val, fuzzy=True)
                        return dt
                    except:
                        continue
                return None

            start_date = find_date_in_row(11)
            end_date = find_date_in_row(12)
            
            date_str = "Unknown Date"
            if start_date and end_date:
                # Format: Jan 26 - Jan 30
                s_str = start_date.strftime("%b %d")
                e_str = end_date.strftime("%b %d")
                date_str = f"{s_str} - {e_str}"
            elif start_date:
                 date_str = start_date.strftime("%b %d")
            elif end_date:
                 date_str = end_date.strftime("%b %d")
            
            extracted_date = date_str
            
            # 4. Clean Data
            print(f"  Metadata Date: {extracted_date}")
            
            # A. Unmerge ALL cells
            merged_ranges = list(ws.merged_cells.ranges)
            for rng in merged_ranges:
                ws.unmerge_cells(str(rng))
                
            # B. Delete Rows 1-14
            ws.delete_rows(1, 14)
            
            # C. Conditional Delete Column D (4th Column)
            # Check header at (1, 4)
            header_d = str(ws.cell(1, 4).value).strip().upper()
            if "BCRM" in header_d:
                 print(f"  Keeping Column D (BCRM found).")
                 # If we keep D, shifts DO NOT happen.
                 offset = 0
            else:
                 # Delete D
                 ws.delete_cols(4)
                 offset = 1 # We deleted 1 column before target area
            
            # D. Delete "Column T" (Original 20)
            # User wants to remove the "Border" column (Original 20).
            # If we deleted D, this is now 19 (S). If we didn't, it is 20 (T).
            # Target Delete Index = 20 - offset
            ws.delete_cols(20 - offset)
            
            # E. Remove Footer ... (Same logic)
            max_r = ws.max_row
            found_footer_at = None
            for r in range(max_r, max(1, max_r - 20), -1):
                val = str(ws.cell(r, 1).value)
                if val and "Number of Record" in val:
                    found_footer_at = r
                    break
            
            if found_footer_at:
                count = max_r - found_footer_at + 1
                if count > 0: ws.delete_rows(found_footer_at, count)
            else:
               if ws.max_row > 2: ws.delete_rows(ws.max_row - 1, 2)

            # F. Resize & Add Image Formulas
            # 1. FIND URL Column
            url_col_idx = 18 - offset # Fallback: Original R (18) shifted by offset
            
            # Dynamic Scan
            for col in range(1, ws.max_column + 1):
                val = str(ws.cell(1, col).value).upper()
                if "URL" in val and "ATTACH" in val:
                    url_col_idx = col
                    break
            
            img_col_idx = url_col_idx + 1
            img_col_letter = get_column_letter(img_col_idx)

            # 2. Add Header
            ws.cell(1, img_col_idx).value = "IMAGES"

            # 3. Resize Rows & Img Column
            ws.column_dimensions[img_col_letter].width = 33
            
            for r in range(2, ws.max_row + 1):
                ws.row_dimensions[r].height = 180
                
                # 4. Insert Formula: using _xlfn.IMAGE for compatibility
                # Formula: =IMAGE(Reference,,1)
                url_ref = f"{get_column_letter(url_col_idx)}{r}"
                ws.cell(r, img_col_idx).value = f"=_xlfn.IMAGE({url_ref},,1)"

            # 5. Save
            clean_filename = f"LKS Data ({extracted_date}).xlsx"
            clean_path = file_path.parent / clean_filename
            wb.save(clean_path)
            
            print(f"  Processed Legacy File -> {clean_filename}")
            return clean_path
            
        finally:
            # Cleanup temp
            if temp_xlsx.exists():
                try: os.remove(temp_xlsx)
                except: pass
