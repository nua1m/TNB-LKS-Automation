# core/image.py — v4.1 (No @, openpyxl-safe)

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

from core.so_utils import clean_so
from config import (
    ATTACH_SHEET_NAME,
    DATA_START_ROW,
    SERVICE_ORDER_COL_IDX,
    COL_ATTACH_URL,
    COL_3MS_SO,
)

# ------------------------------------------------------------
# Strict keyword detection
# ------------------------------------------------------------
def detect_type(url: str) -> str | None:
    if not url:
        return None

    u = url.lower()

    if "old_read" in u:
        return "old"
    if "card" in u:
        return "card"
    if "new_meter" in u:
        return "new"

    return None


# ------------------------------------------------------------
# Build URL mapping
# ------------------------------------------------------------
def build_url_map(df):
    url_map = {}

    for _, row in df.iterrows():
        so = clean_so(row[COL_3MS_SO])
        url = str(row[COL_ATTACH_URL]).strip()

        if not so:
            continue

        if so not in url_map:
            url_map[so] = {"old": None, "card": None, "new": None}

        if not url:
            continue

        t = detect_type(url)

        if t == "old" and url_map[so]["old"] is None:
            url_map[so]["old"] = url
        elif t == "card" and url_map[so]["card"] is None:
            url_map[so]["card"] = url
        elif t == "new" and url_map[so]["new"] is None:
            url_map[so]["new"] = url

    return url_map


# ------------------------------------------------------------
# Clean excel formula
# ------------------------------------------------------------
def img_formula(url: str) -> str | None:
    if not url:
        return None
    return f'=_xlfn.IMAGE("{url}",,1)'


# ------------------------------------------------------------
# SAFE formula writer for openpyxl
# ------------------------------------------------------------
def set_formula_no_at(cell, formula):
    """
    Prevent @IMAGE by forcing Excel to treat the cell as a formula.
    """
    if not formula:
        cell.value = ""
        return

    cell.value = formula
    cell.data_type = "f"  # <-- THIS prevents @ insertion


# ------------------------------------------------------------
# Image Pipeline
# ------------------------------------------------------------
def run_image_pipeline(data_path: str, template_path: str, progress_cb=None):

    df = pd.read_excel(data_path, dtype=str).fillna("")
    df[COL_3MS_SO] = df[COL_3MS_SO].replace("", pd.NA).ffill()

    url_map = build_url_map(df)

    wb = load_workbook(template_path)
    wsA = wb[ATTACH_SHEET_NAME]

    last_row = wsA.max_row
    if last_row < DATA_START_ROW:
        last_row = DATA_START_ROW

    col_old, col_card, col_new = 4, 5, 6
    missing = []
    idx = 0
    total = last_row - DATA_START_ROW + 1

    # -----------------------------------------------------
    # WRITE FORMULAS WITHOUT @
    # -----------------------------------------------------
    for r in range(DATA_START_ROW, last_row + 1):

        so = clean_so(wsA.cell(r, SERVICE_ORDER_COL_IDX).value)
        if not so:
            continue

        idx += 1

        imgs = url_map.get(so, {})
        f_old = img_formula(imgs.get("old"))
        f_card = img_formula(imgs.get("card"))
        f_new = img_formula(imgs.get("new"))

        set_formula_no_at(wsA.cell(r, col_old), f_old)
        set_formula_no_at(wsA.cell(r, col_card), f_card)
        set_formula_no_at(wsA.cell(r, col_new), f_new)

        if progress_cb:
            progress_cb(f"Processing SO {so} ({idx}/{total})")

    wb.save(template_path)

    # -----------------------------------------------------
    # VERIFY MISSING IMAGES
    # -----------------------------------------------------
    wb = load_workbook(template_path, data_only=False)
    wsA = wb[ATTACH_SHEET_NAME]

    for r in range(DATA_START_ROW, last_row + 1):
        so = clean_so(wsA.cell(r, SERVICE_ORDER_COL_IDX).value)
        if not so:
            continue

        if (
            not wsA.cell(r, col_old).value
            or not wsA.cell(r, col_card).value
            or not wsA.cell(r, col_new).value
        ):
            missing.append(so)

    wb.close()
    return sorted(set(missing))
