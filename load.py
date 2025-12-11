# load.py — LKS Builder v4 (FULLY COMPATIBLE WITH IMAGE PIPELINE v4)

import sys
import time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

from ui.ascii import show_title
from ui.layout import set_window_size
from ui.components import summary_block, step_progress
from ui.colors import CYAN, GREEN, YELLOW, RED, RESET, DIM

from core.text import (
    build_claim_rows,
    write_to_claim_sheet,
    write_to_attachment_sheet,
)
from core.image import run_image_pipeline      # <-- v4 pipeline
from core.so_utils import clean_so

from config import CLAIM_SHEET_NAME, ATTACH_SHEET_NAME

RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------
# Helper: detect next empty row in template (Column B)
# ---------------------------------------------------------
def get_next_empty_row(ws, col_idx=2, start_row=3):
    for r in range(start_row, ws.max_row + 1):
        val = ws.cell(row=r, column=col_idx).value
        if val in (None, "", " "):
            return r
    return ws.max_row + 1


# ---------------------------------------------------------
# Helper: center alignment
# ---------------------------------------------------------
def apply_center_formatting(ws, start_row=3):
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = CENTER


# ---------------------------------------------------------
# MAIN
# ---------------------------------------------------------
def main():
    if len(sys.argv) < 3:
        print("Usage: python load.py <data.xlsx> <template.xlsx>")
        sys.exit(1)

    data = Path(sys.argv[1]).resolve()
    template = Path(sys.argv[2]).resolve()

    set_window_size(110, 40)
    show_title()

    print(f"{CYAN}RAW       : {RESET}{data}")
    print(f"{CYAN}TEMPLATE  : {RESET}{template}\n")

    start_time = time.time()

    # -----------------------------------------------------
    # STEP 1 — BUILD CLAIM ROWS
    # -----------------------------------------------------
    print(f"{CYAN}› Reading RAW Data...{RESET}")
    claim_rows, stats = build_claim_rows(data)

    print(f"{DIM}  • SOs after TRAS removal     : {RESET}{GREEN}{stats['sos_after_tras']}{RESET}")
    print(f"{DIM}  • Duplicates skipped in RAW  : {RESET}{GREEN}{stats['duplicates_skipped']}{RESET}")
    print(f"{DIM}  • TRAS removed               : {RESET}{GREEN}{stats['tras_removed']}{RESET}\n")

    # -----------------------------------------------------
    # STEP 2 — PREP TEMPLATE
    # -----------------------------------------------------
    print(f"{CYAN}› Preparing TEMPLATE...{RESET}")

    wb = load_workbook(template)
    wsC = wb[CLAIM_SHEET_NAME]
    wsA = wb[ATTACH_SHEET_NAME]

    # Collect existing SOs
    existing_sos = set()
    for r in range(3, wsC.max_row + 1):
        so = clean_so(wsC.cell(r, 2).value)
        if so:
            existing_sos.add(so)

    new_rows = [row for row in claim_rows if clean_so(row["Service Order"]) not in existing_sos]

    if existing_sos:
        print(f"{YELLOW}The LKS already contains data. Continue adding new SOs? (y/n){RESET}")
        if input(">> ").strip().lower() != "y":
            print(f"{RED}Aborted.{RESET}")
            wb.close()
            return

    if not new_rows:
        print(f"{YELLOW}All SOs already exist in TEMPLATE.{RESET}")
        wb.close()
        return

    start_claim = get_next_empty_row(wsC)
    start_attach = get_next_empty_row(wsA)

    # -----------------------------------------------------
    # STEP 3 — WRITE CLAIM & ATTACHMENT
    # -----------------------------------------------------
    print(f"{CYAN}› Writing CLAIM & ATTACHMENT...{RESET}")

    write_to_claim_sheet(wsC, new_rows, start_row=start_claim)
    write_to_attachment_sheet(wsA, new_rows, start_row=start_attach)

    apply_center_formatting(wsC)
    apply_center_formatting(wsA)

    wb.save(template)
    wb.close()

    print()

    # -----------------------------------------------------
    # STEP 4 — IMAGE PIPELINE v4
    # -----------------------------------------------------
    print("› Processing IMAGES...\n")

    total = len(new_rows)
    prog = 0

    def img_progress(extra=""):
        nonlocal prog
        prog += 1
        step_progress("IMAGES", prog, total, extra=extra, spinner_i=prog)

    # STRICT v4: returns SOs missing ANY of 3 required photos
    defective_sos = run_image_pipeline(str(data), str(template), progress_cb=img_progress)

    print("\n")

    # -----------------------------------------------------
    # STEP 5 — DETAILED MISSING-IMAGE ANALYSIS
    # -----------------------------------------------------
    print(f"{RED}› Analyzing missing image types...{RESET}")

    wb = load_workbook(template, data_only=False)
    wsA = wb[ATTACH_SHEET_NAME]
    wsC = wb[CLAIM_SHEET_NAME]

    missing_detail = {}
    miss_old = miss_card = miss_new = 0

    col_old, col_card, col_new = 4, 5, 6

    for r in range(3, wsA.max_row + 1):
        so = clean_so(wsA.cell(r, 2).value)
        if not so:
            continue

        missing_slots = []

        if not wsA.cell(r, col_old).value:
            missing_slots.append("old_meter")
            miss_old += 1

        if not wsA.cell(r, col_card).value:
            missing_slots.append("card")
            miss_card += 1

        if not wsA.cell(r, col_new).value:
            missing_slots.append("new_meter")
            miss_new += 1

        if missing_slots:
            missing_detail[so] = missing_slots

    # Print missing breakdown
    if missing_detail:
        print(f"{YELLOW}Missing images:{RESET}\n")
        for so, slots in missing_detail.items():
            print(f"  • SO {so} → {', '.join(slots)}")
    else:
        print(f"{GREEN}All SOs have complete images!{RESET}")

    print()

    # -----------------------------------------------------
    # STEP 6 — MARK DEFECTIVE ROWS
    # -----------------------------------------------------
    print("› Marking defective rows...")

    defective_set = set(missing_detail.keys())

    # CLAIM
    for r in range(3, wsC.max_row + 1):
        if clean_so(wsC.cell(r, 2).value) in defective_set:
            for cell in wsC[r]:
                cell.fill = RED_FILL

    # ATTACHMENT
    for r in range(3, wsA.max_row + 1):
        if clean_so(wsA.cell(r, 2).value) in defective_set:
            for cell in wsA[r]:
                cell.fill = RED_FILL

    apply_center_formatting(wsC)
    apply_center_formatting(wsA)

    wb.save(template)
    wb.close()

    # -----------------------------------------------------
    # STEP 7 — SUMMARY
    # -----------------------------------------------------
    elapsed = time.time() - start_time

    summary_block(
        {
            "Total SOs in RAW (after TRAS)": stats["sos_after_tras"],
            "TRAS removed": stats["tras_removed"],
            "Duplicates skipped in RAW": stats["duplicates_skipped"],
            "New SOs appended": len(new_rows),
            "Existing SOs skipped": len(claim_rows) - len(new_rows),
            "Missing OLD meter": miss_old,
            "Missing CARD": miss_card,
            "Missing NEW meter": miss_new,
            "Total defective rows": len(defective_set),
            "Execution time": f"{elapsed:.2f}s",
        },
        str(template),
    )


if __name__ == "__main__":
    main()
