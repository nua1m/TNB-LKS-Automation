import sys
import base64
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from PIL import Image
import requests
import re
from datetime import datetime

# Import DateEngine
from core.services.date_engine import DateEngine

# Column indices (1-based for openpyxl)
COL_SO = 2          # B (Service Order)
COL_DATE = 9        # I (Status Date)
COL_HARI = 10       # J (Hari Field)
COL_REMARKS_1 = 17  # Q (Remarks 1)
COL_REMARKS_2 = 18  # R (Remarks 2)

def extract_date_from_text(text):
    """Extract date from AI response using regex patterns."""
    if not text:
        return None
    
    # Common date patterns
    patterns = [
        r'(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})',  # "4 Dec 2025"
        r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',  # "04/12/2025"
        r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})',  # "2025-12-04"
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            try:
                if 'A-Za-z' in pattern:  # Month name format
                    day, month_name, year = match.groups()
                    date_str = f"{day} {month_name} {year}"
                    return datetime.strptime(date_str, '%d %b %Y')
                else:
                    parts = match.groups()
                    for fmt in ['%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d']:
                        try:
                            return datetime.strptime('/'.join(parts), fmt)
                        except:
                            continue
            except Exception as e:
                print(f"      Failed to parse date from: {text} ({e})")
                continue
    
    return None

def query_moondream(image_path, prompt):
    """Query Moondream via Ollama API."""
    # Read and encode image as base64
    with open(image_path, 'rb') as f:
        image_data = base64.b64encode(f.read()).decode('utf-8')
    
    # Ollama chat API endpoint
    url = "http://localhost:11434/api/chat"
    
    payload = {
        "model": "llava:7b",
        "messages": [
            {
                "role": "user",
                "content": prompt,
                "images": [image_data]
            }
        ],
        "stream": False
    }
    
    response = requests.post(url, json=payload)
    response.raise_for_status()
    
    result = response.json()
    return result['message']['content']

def enhance_excel(excel_path, images_dir):
    """Enhance Excel with Moondream (via Ollama) + DateEngine."""
    
    img_dir = Path(images_dir)
    if not img_dir.exists():
        print(f"ERROR: Image directory not found: {img_dir}")
        return
    
    # Test Ollama connection
    print("Testing Ollama connection...")
    try:
        response = requests.get("http://localhost:11434/api/tags")
        response.raise_for_status()
        models = [m['name'] for m in response.json()['models']]
        if 'llava:7b' not in models and 'llava' not in [m.split(':')[0] for m in models]:
            print("ERROR: Llava model not found. Run: ollama pull llava:7b")
            return
        print("   Ollama connected! Llava ready.")
    except Exception as e:
        print(f"ERROR: Cannot connect to Ollama: {e}")
        print("Make sure Ollama is running in the background.")
        return
    
    # Load workbook
    print(f"Loading Workbook: {excel_path}...")
    wb = load_workbook(excel_path)
    
    if 'CLAIM' not in wb.sheetnames:
        print("'CLAIM' sheet not found!")
        return
    
    ws = wb['CLAIM']
    
    # Yellow fill for Diskon
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    processed_count = 0
    diskon_count = 0
    
    print(f"\nProcessing rows...")
    
    # Start from row 3
    for r in range(3, ws.max_row + 1):
        so_val = ws.cell(r, COL_SO).value
        
        if not so_val:
            continue
        
        so = str(so_val).strip()
        
        # Try to find image (priority: old -> new -> ticket)
        img_path = None
        for suffix in ["old", "new", "ticket"]:
            for ext in [".png", ".jpg", ".jpeg"]:
                p = img_dir / f"{so}_{suffix}{ext}"
                if p.exists():
                    img_path = p
                    break
            if img_path:
                break
        
        if not img_path:
            continue  # Skip if no image found
        
        # Vision AI Query
        try:
            prompt = """Look at the TOP RIGHT corner of this meter image. There should be a date displayed.

Instructions:
1. Find the date in the TOP RIGHT area of the meter display
2. Extract ONLY the date numbers (day, month, year)
3. Return your answer in EXACTLY this format: DD Mon YYYY
   Example: 04 Dec 2025
4. If you cannot find a date, respond with exactly: NO DATE
5. Do NOT include any other text, explanation, or punctuation

Your response:"""
            
            response = query_moondream(img_path, prompt)
            
            # DEBUG: Print raw response
            print(f"   [DEBUG] Row {r} (SO {so}): Moondream response = '{response}'")
            
            # Extract date from response
            if "NO DATE" in response.upper() or not response:
                print(f"   [SKIP] Row {r} (SO {so}): No date in AI response: '{response}'")
                continue
            
            ocr_date = extract_date_from_text(response)
            
            if not ocr_date:
                print(f"   [SKIP] Row {r} (SO {so}): Could not parse date from: '{response}'")
                continue
            
            # Get current status date
            current_status_date = ws.cell(r, COL_DATE).value
            current_status_str = current_status_date.strftime('%b %d, %Y, %I:%M %p') if isinstance(current_status_date, datetime) else str(current_status_date)
            
            # Apply DateEngine
            logic = DateEngine.calculate(
                status_date_str=current_status_str,
                ocr_date_str=ocr_date.strftime('%d %b %Y')
            )
            
            # Update Excel
            ws.cell(r, COL_DATE).value = logic["effective_date"]
            ws.cell(r, COL_HARI).value = logic["hari"]
            ws.cell(r, COL_REMARKS_1).value = logic["remarks_1"]
            ws.cell(r, COL_REMARKS_2).value = logic["remarks_2"]
            
            # Highlight Diskon
            if logic["is_diskon"]:
                ws.cell(r, COL_DATE).fill = yellow_fill
                diskon_count += 1
                print(f"   [DISKON] SO {so}: AI={ocr_date.strftime('%d %b %Y')} vs Status={current_status_str}")
            else:
                print(f"   [OK] SO {so}: Date={ocr_date.strftime('%d %b %Y')}")
            
            processed_count += 1
            
            if processed_count % 10 == 0:
                print(f"   ... Processed {processed_count} images")
        
        except Exception as e:
            print(f"   [ERROR] Row {r} (SO {so}): {e}")
            continue
    
    # Save
    print(f"\nSaving workbook...")
    wb.save(excel_path)
    
    print(f"\nDONE!")
    print(f"   Processed: {processed_count} images")
    print(f"   Diskon cases: {diskon_count}")
    print(f"   Updated: {excel_path}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python enhancer_ollama.py <excel_file> <images_directory>")
        sys.exit(1)
    
    enhance_excel(sys.argv[1], sys.argv[2])
