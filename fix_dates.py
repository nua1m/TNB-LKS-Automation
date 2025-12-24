import sys
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

# Add project root to path
sys.path.append(str(Path(__file__).parent.resolve()))

from config import CLAIM_SHEET_NAME
from core.services.date_engine import DateEngine
from ui.colors import GREEN, RED, RESET, CYAN

def fix_dates(file_path):
    path = Path(file_path).resolve()
    if not path.exists():
        print(f"{RED}Error: File not found: {path}{RESET}")
        return

    print(f"{CYAN}Opening workbook: {path.name}...{RESET}")
    try:
        wb = load_workbook(path, keep_vba=True)
        if CLAIM_SHEET_NAME not in wb.sheetnames:
            print(f"{RED}Error: Sheet '{CLAIM_SHEET_NAME}' not found in workbook.{RESET}")
            return
        
        ws = wb[CLAIM_SHEET_NAME]
        
        # Find "Status Date" column header
        header_row = 1 # Assuming header is in row 1-3. Let's scan first few rows.
        status_date_col = None
        
        # Scan first 5 rows for header
        for r in range(1, 6):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val and str(val).strip() == "Status Date":
                    status_date_col = c
                    header_row = r
                    break
            if status_date_col: break
            
        if not status_date_col:
            print(f"{RED}Error: Could not find 'Status Date' column.{RESET}")
            return

        print(f"Found 'Status Date' at Column {status_date_col} (Row {header_row})")
        
        updates = 0
        total = 0
        
        # Iterate rows
        for row in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_date_col)
            val = cell.value
            
            if val:
                total += 1
                # If already datetime, update format to be safe
                if isinstance(val, datetime):
                    cell.number_format = "d mmm, yyyy, h:mm AM/PM"
                    updates += 1
                else:
                    # Parse string
                    dt = DateEngine.parse_datetime(val)
                    if dt:
                        cell.value = dt
                        cell.number_format = "d mmm, yyyy, h:mm AM/PM"
                        updates += 1
                        
        print(f"{GREEN}Updated {updates}/{total} date cells.{RESET}")
        print("Saving...")
        wb.save(path)
        print(f"{GREEN}Done!{RESET}")

    except Exception as e:
        print(f"{RED}An error occurred: {e}{RESET}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python fix_dates.py <ExcelFile.xlsx>")
    else:
        fix_dates(sys.argv[1])
